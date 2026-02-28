import os
import datetime
from openpyxl import Workbook, load_workbook

# This function creates an Excel file to track the status of downloaded files. It takes the main path, sub-path, pipeline name, company name, tariff program, effective status, file status, and time taken as input. If the tracker file for the current date already exists, it appends a new entry to it; otherwise, it creates a new tracker file with headers and the first entry.
def create_excel_tracker_files(main_path, path, pipelines, company_name, tariff_program, is_effective, file_status, time_taken):
    try:
        tracker_path = os.path.join(main_path, path)
    
        if not os.path.exists(tracker_path):
            os.makedirs(tracker_path)

        today = datetime.date.today()
        tracker_file = os.path.join(tracker_path, f"file_status_record_{today}.xlsx")

        new_entry = [pipelines, company_name, tariff_program, is_effective, file_status, time_taken]

        if os.path.exists(tracker_file):
            wb = load_workbook(tracker_file)
            ws = wb.active
            ws.append(new_entry)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Pipeline', 'Company Name', 'Tariff Title', 'Effective Status', 'File Status', 'Time Taken'])
            ws.append(new_entry)
        
        wb.save(tracker_file)

    except Exception as e:
        print(f"Error creating tracker file: {e}")

    
# This function creates a folder for each pipeline inside the specified main path and sub-path to store the downloaded pdf files.
def create_pipeline_folder(mainpath, path, pipeline_name):
    
    try:
        input_path = os.path.join(mainpath, path)

        if not os.path.exists(input_path):
            os.makedirs(input_path)

        pipeline_folder = os.path.join(input_path, pipeline_name)
        if not os.path.exists(pipeline_folder):
            os.makedirs(pipeline_folder)
            print(f"Folder created for {pipeline_name}")
            return pipeline_folder
        else:
            print(f"Folder already exists for {pipeline_name}") 
            return None

    except Exception as e:
        print(f"Error creating pipeline folder: {e}")

    